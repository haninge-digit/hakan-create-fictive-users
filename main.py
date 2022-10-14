import asyncio
from collections import UserList
import os
import time
import logging
import json

from pymongo import MongoClient
from openpyxl import load_workbook

""" 
Environment
"""
MONGO_URI = os.getenv('MONGO_URI',"mongodb://localhost:27017")          # MONGO_DB adress and port

DEBUG_MODE = os.getenv('DEBUG','false') == "true"                       # Global DEBUG logging

LOGFORMAT = "%(asctime)s %(funcName)-10s [%(levelname)s] %(message)s"   # Log format


def read_values(excel_data):
	# wb = load_workbook(BytesIO(excel_data), read_only=True)
	wb = load_workbook(filename=excel_data, read_only=True)
	ws = wb.worksheets[0]

	db = {}
	headers = {}
	first_row = True
	for row in ws.rows:
		col = 0

		for cell in row:
			if first_row:
				headers[col] = cell.value
			else:
				if cell.value:
					if col == 0:
						person_id = cell.value
						db[person_id] = {'PersonId': person_id, 'children':[]}
					elif 'Child' in headers[col]:
						db[person_id]['children'].append(cell.value)
					else:
						db[person_id][headers[col]] = cell.value
			col += 1

		if first_row:
			first_row = False

	wb.close()         # Close the workbook after reading
	return db

def get_userdata(user):
    data = {
        "PersonId": f" {user['PersonId']} ",
        "Address": f" {user['Address']} ",
        "BirthPlace": "  ",
        "City": f" {user['City']} ",
        "CivilStatus": "  ",
        "Country": "  ",
        "FirstName": f" {user['FirstName']} ",
        "GivenName": f" {user['GivenName']} ",
        "LastName": f" {user['LastName']} ",
        "ZipCode": f" {user['ZipCode']} ",
        "MunicipalityCode": f"{user['MunicipalityCode']} ",
        "Parish": "",
        'Relation': []
    }
    return data

def main():
    users = read_values('Testanvändare.xlsx')
    # print(json.dumps(users, indent=2))

    db = {}
    for user in users:
        if user not in db:
            userdata = get_userdata(users[user])
            for child in users[user]['children']:
                if child not in db:
                    childdata = get_userdata(users[child])
                else:
                    childdata = db[child]
                parent_sex = 'MO' if (int(user) % 2) == 0 else 'FA'
                childdata['Relation'].append({"RelationId": {"PersonNr": user},"Relationstyp": parent_sex})
                db[child] = childdata
                userdata['Relation'].append({"RelationId": {"PersonNr": child},"Relationstyp": 'B'})
            db[user] = userdata

    # print(json.dumps(db, indent=2))
    client = MongoClient(MONGO_URI)
    mongo_db = client.UserInfoCash      # DB is UserInfoCash

    for user_id,user_data in db.items():
        cash_data = mongo_db.external.find_one({'_id':user_id})
        record = {'_id':user_id, 'updated':32536799999, 'data':user_data, 'extra':{}}   # Update time is in the far future = Will always be valid
        if cash_data:
            extra_data = cash_data['extra']
            record['extra'] = extra_data    # Keep old extra data
            res = mongo_db.external.replace_one({'_id':user_id}, record)     # Update existing
        else:
            res = mongo_db.external.insert_one(record)  # Create a new cash record

""" 
Starting point
"""
if __name__ == "__main__":
	if DEBUG_MODE:       # Debug requested
		logging.basicConfig(level=logging.DEBUG, format=LOGFORMAT)
	logging.basicConfig(level=logging.INFO, format=LOGFORMAT)     # Default logging level

	main()